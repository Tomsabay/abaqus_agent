"""Assemble an Excel data workbook (数据册) from N finished run directories.

BUILD_SPEC W20 criterion 2 / 8. Usage:

    .venv\\Scripts\\python.exe reporting\\build_xlsx_workbook.py <run_dir> [<run_dir> ...] \\
        --limits course\\materials\\module5\\limits.yaml \\
        --out    artifacts\\course\\5_3\\数据册.xlsx

Sheets: 汇总 (one data row per run, raw float KPI values so a cell equals the
number in ``_kpi_result.json`` bit for bit) · 限值表 (only with --limits) ·
one detail sheet per run (KPI vs limit vs utilisation + a native bar chart) ·
扫描结果 (only for runs that happen to carry ``parametric_sweep/sweep_results.json``
— the workbook works fine without it).

Over-limit cells in 汇总 are flagged by real conditional-formatting rules that
travel with the file, so the highlight survives in Excel without a macro.

Exit codes: 0 ok · 2 bad input/limits.
"""

from __future__ import annotations

import argparse
import contextlib
import json
import math
import re
import sys
from pathlib import Path
from typing import Any

if __package__ in (None, ""):  # direct `python reporting/build_xlsx_workbook.py`
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook  # noqa: E402
from openpyxl.chart import BarChart, Reference  # noqa: E402
from openpyxl.formatting.rule import CellIsRule  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from reporting.limits import (  # noqa: E402
    NO_LIMITS_NOTICE,
    Judgement,
    LimitsError,
    judge_kpis,
    load_limits,
    overall_verdict,
)
from reporting.run_bundle import (  # noqa: E402
    RunBundle,
    actual_abaqus_release,
    demo_notice,
    generated_at,
    load_run_bundle,
)
from reporting.templates import DEMO_BANNER  # noqa: E402

SUMMARY_SHEET = "汇总"
LIMITS_SHEET = "限值表"
SWEEP_SHEET_PREFIX = "扫描"
OVER_LIMIT_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
BANNER_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FIXED_COLUMNS = ("序号", "案例目录", "模型", "run_id", "状态", "KPI 数")
_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


@contextlib.contextmanager
def exact_float_cells():
    """Write float cells with full precision while inside this context.

    openpyxl serialises numbers with ``"%.16g"`` (openpyxl/compat/strings.py),
    which drops the 17th significant digit and therefore changes the double:
    115.36862182617188 lands in the sheet as 115.3686218261719. This workbook is
    an evidence artifact whose cells must equal ``_kpi_result.json`` digit for
    digit, so swap in ``repr`` for finite floats during save only.
    """
    from openpyxl.cell import _writer as cell_writer

    original = cell_writer.safe_string

    def exact(value):
        if isinstance(value, float) and math.isfinite(value):
            return repr(value)
        return original(value)

    cell_writer.safe_string = exact
    try:
        yield
    finally:
        cell_writer.safe_string = original


def _style_header(worksheet, row: int, columns: int) -> None:
    for column in range(1, columns + 1):
        cell = worksheet.cell(row=row, column=column)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(worksheet, widths: dict[int, int]) -> None:
    for column, width in widths.items():
        worksheet.column_dimensions[get_column_letter(column)].width = width


def _sheet_title(index: int, bundle: RunBundle, used: set[str]) -> str:
    base = _INVALID_SHEET_CHARS.sub("_", f"{index}_{bundle.model_name}")[:31] or f"run{index}"
    title = base
    suffix = 1
    while title in used:
        suffix += 1
        title = f"{base[:28]}_{suffix}"
    used.add(title)
    return title


def _kpi_columns(bundles: list[RunBundle]) -> list[str]:
    """Union of KPI names, first-seen order (stable across reruns)."""
    names: list[str] = []
    for bundle in bundles:
        for name in bundle.kpis:
            if name not in names:
                names.append(name)
    return names


def _write_summary(worksheet, bundles: list[RunBundle], kpi_names: list[str],
                   limits: dict, release: str) -> dict[str, Any]:
    demo_present = any(b.demo_mode for b in bundles)
    row = 1
    if demo_present:
        notice = next((demo_notice(b) for b in bundles if b.demo_mode), "")
        cell = worksheet.cell(row=1, column=1, value=f"{DEMO_BANNER} —— {notice}")
        cell.font = Font(bold=True)
        cell.fill = BANNER_FILL
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1,
                             end_column=max(len(FIXED_COLUMNS) + len(kpi_names), 2))
        row = 2

    header_row = row
    header = list(FIXED_COLUMNS) + kpi_names
    for column, title in enumerate(header, start=1):
        worksheet.cell(row=header_row, column=column, value=title)
    _style_header(worksheet, header_row, len(header))
    worksheet.freeze_panes = worksheet.cell(row=header_row + 1, column=1)

    first_data_row = header_row + 1
    for index, bundle in enumerate(bundles, start=1):
        excel_row = header_row + index
        values = [
            index,
            str(bundle.run_dir),
            bundle.model_name,
            bundle.run_id,
            bundle.status,
            len(bundle.kpis),
        ]
        for column, value in enumerate(values, start=1):
            worksheet.cell(row=excel_row, column=column, value=value)
        for offset, name in enumerate(kpi_names):
            value = bundle.kpis.get(name)
            if value is None:
                continue
            worksheet.cell(row=excel_row, column=len(FIXED_COLUMNS) + 1 + offset,
                           value=float(value) if isinstance(value, (int, float)) else str(value))
    last_data_row = header_row + len(bundles)

    footer_row = last_data_row + 2
    worksheet.cell(row=footer_row, column=1,
                   value=f"生成时间 {generated_at()}；Abaqus（实测）{release}；"
                         f"数值直接取自各 run 的 _kpi_result.json，未做任何四舍五入")

    rules = _apply_conditional_formatting(
        worksheet, kpi_names, limits, first_data_row, last_data_row,
    )
    _autosize(worksheet, {1: 6, 2: 46, 3: 18, 4: 20, 5: 12, 6: 8})
    for offset in range(len(kpi_names)):
        _autosize(worksheet, {len(FIXED_COLUMNS) + 1 + offset: 18})
    return {
        "header_row": header_row,
        "first_data_row": first_data_row,
        "last_data_row": last_data_row,
        "data_rows": len(bundles),
        "kpi_columns": kpi_names,
        "conditional_formatting_rules": rules,
        "demo_banner_row": 1 if demo_present else 0,
    }


def _apply_conditional_formatting(worksheet, kpi_names: list[str], limits: dict,
                                  first_row: int, last_row: int) -> list[dict[str, Any]]:
    """Real CF rules (not static fills) on every KPI column that has a limit."""
    applied: list[dict[str, Any]] = []
    if last_row < first_row:
        return applied
    for offset, name in enumerate(kpi_names):
        limit = limits.get(name)
        if limit is None:
            continue
        column_letter = get_column_letter(len(FIXED_COLUMNS) + 1 + offset)
        cell_range = f"{column_letter}{first_row}:{column_letter}{last_row}"
        operator = "greaterThan" if limit.direction == "max" else "lessThan"
        worksheet.conditional_formatting.add(
            cell_range,
            CellIsRule(operator=operator, formula=[repr(limit.value)], fill=OVER_LIMIT_FILL),
        )
        applied.append({
            "kpi": name,
            "range": cell_range,
            "operator": operator,
            "limit": limit.value,
            "unit": limit.unit,
        })
    return applied


def _write_limits_sheet(worksheet, limits: dict, limits_path: str) -> None:
    header = ["KPI", "说明", "限值", "单位", "方向", "关系", "取绝对值", "限值来源"]
    for column, title in enumerate(header, start=1):
        worksheet.cell(row=1, column=column, value=title)
    _style_header(worksheet, 1, len(header))
    for index, name in enumerate(sorted(limits), start=1):
        limit = limits[name]
        for column, value in enumerate(
            [name, limit.label, limit.value, limit.unit, limit.direction,
             limit.relation, "是" if limit.absolute else "否", limit.source],
            start=1,
        ):
            worksheet.cell(row=index + 1, column=column, value=value)
    worksheet.cell(row=len(limits) + 3, column=1,
                   value=f"限值文件：{limits_path}（判定权归本文件，脚本只做算术比较）")
    _autosize(worksheet, {1: 20, 2: 22, 3: 14, 4: 10, 5: 8, 6: 6, 7: 10, 8: 90})


def _write_run_sheet(worksheet, bundle: RunBundle, judgements: list[Judgement],
                     limits_path: str | None) -> dict[str, Any]:
    worksheet.cell(row=1, column=1, value=f"{bundle.model_name}  run_id={bundle.run_id}")
    worksheet.cell(row=1, column=1).font = Font(bold=True)
    worksheet.cell(row=2, column=1, value=f"run 目录：{bundle.run_dir}")
    worksheet.cell(row=3, column=1, value=f"状态：{bundle.status}；KPI 来源：{bundle.kpi_source or '-'}")

    if bundle.demo_mode or not bundle.kpis:
        cell = worksheet.cell(row=5, column=1, value=f"{DEMO_BANNER} —— {demo_notice(bundle)}")
        cell.font = Font(bold=True)
        cell.fill = BANNER_FILL
        _autosize(worksheet, {1: 70})
        return {"sheet": worksheet.title, "chart": False, "rows": 0}

    header_row = 5
    header = ["KPI", "说明", "数值", "单位", "限值", "关系", "利用率(%)", "判定", "限值来源"]
    for column, title in enumerate(header, start=1):
        worksheet.cell(row=header_row, column=column, value=title)
    _style_header(worksheet, header_row, len(header))

    for index, j in enumerate(judgements, start=1):
        row = header_row + index
        worksheet.cell(row=row, column=1, value=j.name)
        worksheet.cell(row=row, column=2, value=j.label or "-")
        worksheet.cell(row=row, column=3, value=float(j.actual))
        worksheet.cell(row=row, column=4, value=j.unit or "-")
        if j.limit is None:
            worksheet.cell(row=row, column=5, value="-")
            worksheet.cell(row=row, column=6, value="-")
            worksheet.cell(row=row, column=7, value="-")
            worksheet.cell(row=row, column=8, value=j.verdict)
            worksheet.cell(row=row, column=9, value="-")
            continue
        worksheet.cell(row=row, column=5, value=float(j.limit.value))
        worksheet.cell(row=row, column=6, value=j.limit.relation)
        worksheet.cell(row=row, column=7, value=float(j.utilisation_pct or 0.0))
        verdict_cell = worksheet.cell(row=row, column=8, value=j.verdict)
        if not j.ok:
            verdict_cell.fill = OVER_LIMIT_FILL
            verdict_cell.font = Font(bold=True)
        worksheet.cell(row=row, column=9, value=j.source)

    last_row = header_row + len(judgements)
    judged = [j for j in judgements if j.judged]
    chart_added = False
    if judged:
        chart = BarChart()
        chart.type = "col"
        chart.title = f"{bundle.model_name} 各 KPI 利用率(%)"
        chart.y_axis.title = "利用率 %"
        chart.x_axis.title = "KPI"
        data = Reference(worksheet, min_col=7, min_row=header_row, max_row=last_row)
        categories = Reference(worksheet, min_col=1, min_row=header_row + 1, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 8
        chart.width = 18
        worksheet.add_chart(chart, f"A{last_row + 3}")
        chart_added = True

    note_row = last_row + 1
    if limits_path:
        worksheet.cell(row=note_row, column=1,
                       value=f"限值文件：{limits_path}；总体判定：{overall_verdict(judgements)}")
    else:
        worksheet.cell(row=note_row, column=1, value=NO_LIMITS_NOTICE)
    _autosize(worksheet, {1: 20, 2: 22, 3: 16, 4: 10, 5: 14, 6: 6, 7: 12, 8: 10, 9: 80})
    return {"sheet": worksheet.title, "chart": chart_added, "rows": len(judgements)}


def _sweep_path(bundle: RunBundle) -> Path | None:
    candidate = bundle.run_dir / "parametric_sweep" / "sweep_results.json"
    return candidate if candidate.is_file() else None


SWEEP_FIXED = ("变体", "参数取值", "状态", "缓存命中", "build_s", "queue_s", "solve_s", "post_s")
SWEEP_TIMING_KEYS = ("build_seconds", "queue_seconds", "solve_seconds", "post_seconds")


def _write_sweep_sheet(worksheet, sweep: dict[str, Any]) -> int:
    """Optional sheet for runs that carry a W21 parametric_sweep/sweep_results.json."""
    results = sweep.get("results", []) or []
    kpi_names: list[str] = []
    for item in results:
        for name in (item.get("kpis", {}) or {}):
            if name not in kpi_names:
                kpi_names.append(name)
    header = list(SWEEP_FIXED) + kpi_names
    for column, title in enumerate(header, start=1):
        worksheet.cell(row=1, column=column, value=title)
    _style_header(worksheet, 1, len(header))
    for row_index, item in enumerate(results, start=1):
        row = row_index + 1
        sample = item.get("sample", {}) or {}
        worksheet.cell(row=row, column=1, value="#%s" % item.get("index", row_index - 1))
        worksheet.cell(row=row, column=2,
                       value="、".join(f"{k}={v}" for k, v in sample.items()) or "-")
        worksheet.cell(row=row, column=3, value=str(item.get("status", "-")))
        worksheet.cell(row=row, column=4, value="是" if item.get("cached") else "否")
        for offset, key in enumerate(SWEEP_TIMING_KEYS):
            value = item.get(key)
            if isinstance(value, (int, float)):
                worksheet.cell(row=row, column=5 + offset, value=float(value))
        for offset, name in enumerate(kpi_names):
            value = (item.get("kpis", {}) or {}).get(name)
            if isinstance(value, (int, float)):
                worksheet.cell(row=row, column=len(SWEEP_FIXED) + 1 + offset, value=float(value))
    footer = len(results) + 3
    worksheet.cell(row=footer, column=1,
                   value=f"strategy={sweep.get('strategy')}  "
                         f"max_parallel={sweep.get('max_parallel')}  "
                         f"solve_slots={sweep.get('solve_slots')}  "
                         f"wall_clock_seconds={sweep.get('wall_clock_seconds')}  "
                         f"peak_concurrent_solves={sweep.get('peak_concurrent_solves')}")
    _autosize(worksheet, {1: 8, 2: 28, 3: 14, 4: 10, 5: 10, 6: 10, 7: 10, 8: 10})
    return len(results)


def build_workbook(
    run_dirs: list[str | Path],
    out: str | Path,
    *,
    limits_path: str | Path | None = None,
    release_timeout: float = 90.0,
) -> dict[str, Any]:
    """Build the workbook for N run directories. Returns a summary dict."""
    if not run_dirs:
        raise ValueError("至少需要一个 run 目录")
    bundles = [load_run_bundle(run_dir) for run_dir in run_dirs]
    limits = load_limits(limits_path) if limits_path else {}
    release = actual_abaqus_release(timeout=release_timeout)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = SUMMARY_SHEET
    kpi_names = _kpi_columns(bundles)
    summary_info = _write_summary(summary_sheet, bundles, kpi_names, limits, release)

    if limits:
        _write_limits_sheet(workbook.create_sheet(LIMITS_SHEET), limits, str(limits_path))

    used_titles = {SUMMARY_SHEET, LIMITS_SHEET}
    run_sheets = []
    judgement_map: dict[str, list[Judgement]] = {}
    for index, bundle in enumerate(bundles, start=1):
        judgements = judge_kpis(bundle.kpis, limits) if bundle.kpis else []
        judgement_map[bundle.run_id] = judgements
        sheet = workbook.create_sheet(_sheet_title(index, bundle, used_titles))
        run_sheets.append(_write_run_sheet(
            sheet, bundle, judgements, str(limits_path) if limits_path else None,
        ))
        sweep_file = _sweep_path(bundle)
        if sweep_file is not None:
            try:
                sweep = json.loads(sweep_file.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                sweep = {}
            if isinstance(sweep, dict) and sweep.get("results"):
                sweep_sheet = workbook.create_sheet(f"{SWEEP_SHEET_PREFIX}{index}"[:31])
                used_titles.add(sweep_sheet.title)
                _write_sweep_sheet(sweep_sheet, sweep)

    out_path = Path(out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with exact_float_cells():
        workbook.save(str(out_path))

    return {
        "output_path": str(out_path),
        "bytes": out_path.stat().st_size,
        "run_count": len(bundles),
        "abaqus_release": release,
        "limits_path": str(limits_path) if limits_path else "",
        "sheets": workbook.sheetnames,
        "summary": summary_info,
        "run_sheets": run_sheets,
        "runs": [
            {
                "run_dir": str(b.run_dir),
                "run_id": b.run_id,
                "model_name": b.model_name,
                "status": b.status,
                "demo_mode": b.demo_mode,
                "kpi_source": b.kpi_source,
                "kpis": b.kpis,
                "overall_verdict": overall_verdict(judgement_map[b.run_id]) if limits else "",
                "over_limit": [
                    j.name for j in judgement_map[b.run_id] if j.judged and not j.ok
                ],
            }
            for b in bundles
        ],
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="build_xlsx_workbook.py",
        description="Assemble an Excel data workbook from N finished run directories",
    )
    parser.add_argument("run_dirs", nargs="+", help="One or more run directories")
    parser.add_argument("--out", required=True, help="Output .xlsx path")
    parser.add_argument("--limits", help="Limit table YAML (adds 限值表 + 判定列 + 条件格式)")
    parser.add_argument("--release-timeout", type=float, default=90.0)
    parser.add_argument("--json", action="store_true", dest="as_json", help="Print JSON summary")
    args = parser.parse_args(argv)

    try:
        summary = build_workbook(
            args.run_dirs,
            args.out,
            limits_path=args.limits,
            release_timeout=args.release_timeout,
        )
    except (FileNotFoundError, LimitsError, ValueError) as e:
        print(f"错误：{e}", file=sys.stderr)
        return 2

    if args.as_json:
        print(json.dumps(summary, ensure_ascii=False, indent=2))
    else:
        print(f"已生成数据册：{summary['output_path']}（{summary['bytes']} 字节）")
        print(f"run 数：{summary['run_count']}  工作表：{', '.join(summary['sheets'])}")
        print(f"汇总表数据行：{summary['summary']['data_rows']}"
              f"  条件格式规则：{len(summary['summary']['conditional_formatting_rules'])}")
        for run in summary["runs"]:
            over = "、".join(run["over_limit"]) or "-"
            print(f"  {run['model_name']}({run['run_id']}) KPI={len(run['kpis'])}"
                  f" 判定={run['overall_verdict'] or '-'} 超限={over}")
        if not summary["limits_path"]:
            print(NO_LIMITS_NOTICE)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
